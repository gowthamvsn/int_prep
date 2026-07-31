import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

BOLD = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="538A00")
TOTAL_FILL = PatternFill("solid", fgColor="E5EFD2")
INPUT_FILL = PatternFill("solid", fgColor="EEF2E8")
HEADER_FILL = PatternFill("solid", fgColor="DBE1D3")
thin = Side(style="thin", color="C8CFC0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
NUMFMT = "#,##0"

def section(ws, row, col, text, span=4):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    c = ws.cell(row=row, column=col, value=text)
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    for cc in range(col, col + span):
        ws.cell(row=row, column=cc).fill = SECTION_FILL

def header_row(ws, row, col, labels):
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=col + i, value=lab)
        c.font = BOLD
        c.fill = HEADER_FILL
        c.border = BORDER

def cell(ws, row, col, value, fmt=None, bold=False, fill=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    if fmt:
        c.number_format = fmt
    if bold:
        c.font = BOLD
    if fill:
        c.fill = fill
    if border:
        c.border = BORDER
    return c

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# SHEET 1 -- Llama 2 7B, full stage-by-stage breakdown
# ============================================================
ws = wb.active
ws.title = "7B Breakdown"
autosize(ws, [32, 30, 34, 20])

ws.cell(row=1, column=1, value="Llama 2 7B — Trainable Parameters, Stage by Stage").font = TITLE_FONT
ws.cell(row=2, column=1, value="Real published architecture config. Every value below is a live formula — change the inputs and everything recomputes.").font = Font(italic=True, size=9, color="5B6557")

# --- Inputs ---
section(ws, 4, 1, "MODEL CONFIGURATION (inputs — edit these)")
header_row(ws, 5, 1, ["Parameter", "Value", "Meaning"])
cell(ws, 6, 1, "hidden_size"); cell(ws, 6, 2, 4096, NUMFMT, fill=INPUT_FILL); cell(ws, 6, 3, "width of every token's vector (d_model)")
cell(ws, 7, 1, "n_layers");    cell(ws, 7, 2, 32, NUMFMT, fill=INPUT_FILL);   cell(ws, 7, 3, "number of stacked transformer layers")
cell(ws, 8, 1, "n_heads");     cell(ws, 8, 2, 32, NUMFMT, fill=INPUT_FILL);   cell(ws, 8, 3, "number of parallel attention heads")
cell(ws, 9, 1, "head_dim");    cell(ws, 9, 2, "=B6/B8", NUMFMT);              cell(ws, 9, 3, "hidden_size / n_heads")
cell(ws, 10, 1, "intermediate_size"); cell(ws, 10, 2, 11008, NUMFMT, fill=INPUT_FILL); cell(ws, 10, 3, "FFN's internal expanded width")
cell(ws, 11, 1, "vocab_size"); cell(ws, 11, 2, 32000, NUMFMT, fill=INPUT_FILL); cell(ws, 11, 3, "number of distinct tokens the model knows")

H, L, HD, NH, I, V = "$B$6", "$B$7", "$B$9", "$B$8", "$B$10", "$B$11"

# --- Stage 1: inside one layer ---
section(ws, 13, 1, "STAGE 1 — INSIDE ONE LAYER (of the 32)")
header_row(ws, 14, 1, ["Component", "Shape", "Formula (Excel)", "Parameters"])
cell(ws, 15, 1, "Attention  (Q, K, V, O projections)")
cell(ws, 15, 2, "[4 × hidden × hidden]")
cell(ws, 15, 3, "4 × hidden × hidden")
cell(ws, 15, 4, f"=4*{H}*{H}", NUMFMT)

cell(ws, 16, 1, "FFN  (SwiGLU: gate_proj, up_proj, down_proj)")
cell(ws, 16, 2, "[3 × hidden × intermediate]")
cell(ws, 16, 3, "3 × hidden × intermediate")
cell(ws, 16, 4, f"=3*{H}*{I}", NUMFMT)

cell(ws, 17, 1, "Norms  (2× RMSNorm scale vectors)")
cell(ws, 17, 2, "[2 × hidden]")
cell(ws, 17, 3, "2 × hidden")
cell(ws, 17, 4, f"=2*{H}", NUMFMT)

cell(ws, 18, 1, "ONE LAYER — TOTAL", bold=True, fill=TOTAL_FILL)
cell(ws, 18, 2, "attention + FFN + norms", fill=TOTAL_FILL)
cell(ws, 18, 3, "=SUM(D15:D17)", fill=TOTAL_FILL)
cell(ws, 18, 4, "=SUM(D15:D17)", NUMFMT, bold=True, fill=TOTAL_FILL)

# --- Stage 2: full model, the addition ---
section(ws, 20, 1, "STAGE 2 — FULL MODEL (the addition, running total)")
header_row(ws, 21, 1, ["Component", "Shape", "This stage", "Running total"])

cell(ws, 22, 1, "Token Embedding")
cell(ws, 22, 2, "[vocab × hidden]")
cell(ws, 22, 3, f"={V}*{H}", NUMFMT)
cell(ws, 22, 4, "=C22", NUMFMT)

cell(ws, 23, 1, "All 32 Layers  (n_layers × one-layer-total)")
cell(ws, 23, 2, "[n_layers × 202,383,360]")
cell(ws, 23, 3, f"={L}*D18", NUMFMT)
cell(ws, 23, 4, "=D23+C23", NUMFMT)

cell(ws, 24, 1, "Final Norm")
cell(ws, 24, 2, "[hidden]")
cell(ws, 24, 3, f"={H}", NUMFMT)
cell(ws, 24, 4, "=D24+C24", NUMFMT)

cell(ws, 25, 1, "Output Head  (LM head, untied from input embedding)")
cell(ws, 25, 2, "[vocab × hidden]")
cell(ws, 25, 3, f"={V}*{H}", NUMFMT)
cell(ws, 25, 4, "=D25+C25", NUMFMT)

cell(ws, 26, 1, "TOTAL TRAINABLE PARAMETERS", bold=True, fill=TOTAL_FILL)
cell(ws, 26, 2, "embedding + layers + norm + head", fill=TOTAL_FILL)
cell(ws, 26, 3, "=SUM(C22:C25)", NUMFMT, bold=True, fill=TOTAL_FILL)
cell(ws, 26, 4, "=D25", NUMFMT, bold=True, fill=TOTAL_FILL)

cell(ws, 27, 1, "  → in billions")
cell(ws, 27, 2, "")
cell(ws, 27, 3, "=C26/1000000000")
ws.cell(row=27, column=3).number_format = "0.000"

# --- % breakdown ---
section(ws, 29, 1, "WHERE THE PARAMETERS ACTUALLY LIVE (% of total)")
header_row(ws, 30, 1, ["Component", "Parameters", "% of total"])
cell(ws, 31, 1, "FFN  (all 32 layers)")
cell(ws, 31, 2, f"={L}*D16", NUMFMT)
cell(ws, 31, 3, "=B31/$C$26"); ws.cell(row=31, column=3).number_format = "0.0%"
cell(ws, 32, 1, "Attention  (all 32 layers)")
cell(ws, 32, 2, f"={L}*D15", NUMFMT)
cell(ws, 32, 3, "=B32/$C$26"); ws.cell(row=32, column=3).number_format = "0.0%"
cell(ws, 33, 1, "Embeddings  (input + output head)")
cell(ws, 33, 2, "=C22+C25", NUMFMT)
cell(ws, 33, 3, "=B33/$C$26"); ws.cell(row=33, column=3).number_format = "0.0%"

ws.freeze_panes = "A5"


# ============================================================
# SHEET 2 -- Llama 2 13B, cross-check (proves it's not a coincidence)
# ============================================================
ws2 = wb.create_sheet("13B Cross-check")
autosize(ws2, [32, 30, 34, 20])
ws2.cell(row=1, column=1, value="Llama 2 13B — Same Formula, Different Model").font = TITLE_FONT
ws2.cell(row=2, column=1, value="Identical structure to the 7B sheet. If the total below independently lands near 13B, the formula is confirmed, not tuned to one example.").font = Font(italic=True, size=9, color="5B6557")

section(ws2, 4, 1, "MODEL CONFIGURATION (real published values)")
header_row(ws2, 5, 1, ["Parameter", "Value", "Meaning"])
cell(ws2, 6, 1, "hidden_size"); cell(ws2, 6, 2, 5120, NUMFMT, fill=INPUT_FILL)
cell(ws2, 7, 1, "n_layers");    cell(ws2, 7, 2, 40, NUMFMT, fill=INPUT_FILL)
cell(ws2, 8, 1, "n_heads");     cell(ws2, 8, 2, 40, NUMFMT, fill=INPUT_FILL)
cell(ws2, 9, 1, "head_dim");    cell(ws2, 9, 2, "=B6/B8", NUMFMT)
cell(ws2, 10, 1, "intermediate_size"); cell(ws2, 10, 2, 13824, NUMFMT, fill=INPUT_FILL)
cell(ws2, 11, 1, "vocab_size"); cell(ws2, 11, 2, 32000, NUMFMT, fill=INPUT_FILL)

H2, L2, I2, V2 = "$B$6", "$B$7", "$B$10", "$B$11"

section(ws2, 13, 1, "STAGE 1 — INSIDE ONE LAYER")
header_row(ws2, 14, 1, ["Component", "Shape", "Formula (Excel)", "Parameters"])
cell(ws2, 15, 1, "Attention  (Q, K, V, O)"); cell(ws2, 15, 2, "[4 × hidden × hidden]"); cell(ws2, 15, 3, "4 × hidden × hidden"); cell(ws2, 15, 4, f"=4*{H2}*{H2}", NUMFMT)
cell(ws2, 16, 1, "FFN  (SwiGLU)"); cell(ws2, 16, 2, "[3 × hidden × intermediate]"); cell(ws2, 16, 3, "3 × hidden × intermediate"); cell(ws2, 16, 4, f"=3*{H2}*{I2}", NUMFMT)
cell(ws2, 17, 1, "Norms"); cell(ws2, 17, 2, "[2 × hidden]"); cell(ws2, 17, 3, "2 × hidden"); cell(ws2, 17, 4, f"=2*{H2}", NUMFMT)
cell(ws2, 18, 1, "ONE LAYER — TOTAL", bold=True, fill=TOTAL_FILL); cell(ws2, 18, 2, "", fill=TOTAL_FILL); cell(ws2, 18, 3, "=SUM(D15:D17)", fill=TOTAL_FILL); cell(ws2, 18, 4, "=SUM(D15:D17)", NUMFMT, bold=True, fill=TOTAL_FILL)

section(ws2, 20, 1, "STAGE 2 — FULL MODEL (the addition, running total)")
header_row(ws2, 21, 1, ["Component", "Shape", "This stage", "Running total"])
cell(ws2, 22, 1, "Token Embedding"); cell(ws2, 22, 2, "[vocab × hidden]"); cell(ws2, 22, 3, f"={V2}*{H2}", NUMFMT); cell(ws2, 22, 4, "=C22", NUMFMT)
cell(ws2, 23, 1, "All 40 Layers"); cell(ws2, 23, 2, "[n_layers × one-layer-total]"); cell(ws2, 23, 3, f"={L2}*D18", NUMFMT); cell(ws2, 23, 4, "=D23+C23", NUMFMT)
cell(ws2, 24, 1, "Final Norm"); cell(ws2, 24, 2, "[hidden]"); cell(ws2, 24, 3, f"={H2}", NUMFMT); cell(ws2, 24, 4, "=D24+C24", NUMFMT)
cell(ws2, 25, 1, "Output Head"); cell(ws2, 25, 2, "[vocab × hidden]"); cell(ws2, 25, 3, f"={V2}*{H2}", NUMFMT); cell(ws2, 25, 4, "=D25+C25", NUMFMT)
cell(ws2, 26, 1, "TOTAL TRAINABLE PARAMETERS", bold=True, fill=TOTAL_FILL); cell(ws2, 26, 2, "", fill=TOTAL_FILL); cell(ws2, 26, 3, "=SUM(C22:C25)", NUMFMT, bold=True, fill=TOTAL_FILL); cell(ws2, 26, 4, "=D25", NUMFMT, bold=True, fill=TOTAL_FILL)
cell(ws2, 27, 1, "  → in billions"); cell(ws2, 27, 3, "=C26/1000000000"); ws2.cell(row=27, column=3).number_format = "0.000"
ws2.freeze_panes = "A5"


# ============================================================
# SHEET 3 -- Tokens & Chinchilla scaling
# ============================================================
ws3 = wb.create_sheet("Tokens & Scaling")
autosize(ws3, [40, 26, 40])
ws3.cell(row=1, column=1, value="Parameters vs. Training Tokens — Chinchilla Scaling").font = TITLE_FONT
ws3.cell(row=2, column=1, value="Live reference to the 7B Breakdown sheet's total.").font = Font(italic=True, size=9, color="5B6557")

section(ws3, 4, 1, "CHINCHILLA COMPUTE-OPTIMAL TOKEN COUNT", span=3)
header_row(ws3, 5, 1, ["Quantity", "Value", "Formula"])
cell(ws3, 6, 1, "7B model total parameters (N)"); cell(ws3, 6, 2, "='7B Breakdown'!C26", NUMFMT); cell(ws3, 6, 3, "from '7B Breakdown' sheet")
cell(ws3, 7, 1, "Chinchilla-optimal tokens (D ≈ 20 × N)"); cell(ws3, 7, 2, "=20*B6", NUMFMT); cell(ws3, 7, 3, "20 × N")
cell(ws3, 8, 1, "Real published Llama 2 7B training tokens"); cell(ws3, 8, 2, 2000000000000, NUMFMT, fill=INPUT_FILL); cell(ws3, 8, 3, "2 trillion, from the Llama 2 paper")
cell(ws3, 9, 1, "Actual tokens-per-parameter ratio"); cell(ws3, 9, 2, "=B8/B6"); ws3.cell(row=9,column=2).number_format="0.0"
cell(ws3, 10, 1, "How far beyond Chinchilla-optimal"); cell(ws3, 10, 2, "=B8/B7"); ws3.cell(row=10,column=2).number_format="0.00\"×\""

ws3.cell(row=12, column=1, value="Why: Chinchilla's D≈20N minimizes TRAINING compute for a target loss. It says nothing about INFERENCE cost. A smaller model trained far past the Chinchilla point can match a larger Chinchilla-optimal model's quality while being far cheaper to serve at inference time, forever, for every query — and inference cost dominates a deployed model's real lifetime cost.").alignment = Alignment(wrap_text=True)
ws3.merge_cells("A12:C12")
ws3.row_dimensions[12].height = 60


# ============================================================
# SHEET 4 -- Notes & caveats
# ============================================================
ws4 = wb.create_sheet("Notes & Caveats")
autosize(ws4, [95])
notes = [
    ("SOURCE", "All configuration numbers (hidden_size, n_layers, n_heads, intermediate_size, vocab_size) are Meta's real published Llama 2 config values — not estimates."),
    ("", ""),
    ("WHY THE FORMULA WORKS FOR 7B AND 13B", "Both use standard multi-head attention: every query head has its own key/value head, so Q, K, V, and O are all full [hidden × hidden] matrices — exactly what 4×hidden² assumes."),
    ("", ""),
    ("WHY IT DOESN'T WORK FOR 70B", "Llama 2 70B uses grouped-query attention (GQA): multiple query heads share a smaller number of key/value heads, so K and V are smaller than [hidden × hidden]. Applying this sheet's formula to 70B's config overcounts by roughly 14% (≈78.4B computed vs. the real ≈68.9B). A correct GQA formula needs separate, smaller terms for K and V."),
    ("", ""),
    ("WHY SWIGLU'S INTERMEDIATE SIZE IS ≈2.688× HIDDEN, NOT 4×", "A classic 2-matrix GELU FFN (up + down) at the standard 4× multiplier costs 2×hidden×(4×hidden) = 8×hidden² parameters. SwiGLU uses 3 matrices (gate, up, down) instead of 2 — solving 3×hidden×intermediate = 8×hidden² for intermediate gives (8/3)×hidden ≈ 2.667×hidden. For hidden=4096 that's ≈10,923 — and the real published intermediate_size, 11,008, is that exact target rounded up to the nearest multiple of 256 (11,008 / 256 = 43 exactly), for hardware efficiency."),
    ("", ""),
    ("EMBEDDING TYING", "Llama does NOT tie the input token embedding and the output LM head — they are two separate [vocab × hidden] matrices, both counted separately in this sheet. Some other model families (e.g. GPT-2) DO tie them, which would remove one entire [vocab × hidden] term from the total."),
]
r = 1
ws4.cell(row=r, column=1, value="Notes & Caveats").font = TITLE_FONT
r += 2
for head, body in notes:
    if head:
        c = ws4.cell(row=r, column=1, value=head)
        c.font = BOLD
        r += 1
    if body:
        c = ws4.cell(row=r, column=1, value=body)
        c.alignment = Alignment(wrap_text=True)
        ws4.row_dimensions[r].height = 45
        r += 1
    r += 1

# ============================================================
# SHEET 5 -- Grassroots, REAL scale: every one of the 32 layers and
# 32 heads enumerated individually (no dummy toy values, no "x32" shorthand).
# Real trained weight VALUES aren't public, so this shows the real SHAPE
# at every matmul plus the real parameter-count formula it implies --
# summed layer by layer, head by head, until the running total IS 7B.
# ============================================================
ws5 = wb.create_sheet("Grassroots — All Layers & Heads")
autosize(ws5, [34, 30, 26, 20, 90])
WRAP = Alignment(wrap_text=True, vertical="top")

def why(ws, row, col, text, height=60):
    c = ws.cell(row=row, column=col, value=text)
    c.alignment = WRAP
    c.border = BORDER
    if ws.row_dimensions[row].height is None or ws.row_dimensions[row].height < height:
        ws.row_dimensions[row].height = height
    return c

ws5.cell(row=1, column=1, value="Grassroots — Real Shapes, Every Layer, Every Head").font = TITLE_FONT
ws5.cell(row=2, column=1, value="Real Llama 2 7B dimensions throughout (32,000 vocab, not a toy example). Every row shows the real matrix SHAPE and the parameter-count FORMULA it implies, summed until it IS the 7B total — not compressed with a ×32 shortcut. Column E explains WHY each row is shaped the way it is.").font = Font(italic=True, size=9, color="5B6557")

section(ws5, 4, 1, "MODEL DIMENSIONS (linked live to the '7B Breakdown' sheet — change it there, this sheet follows)", span=5)
header_row(ws5, 5, 1, ["Parameter", "Value", "", "", "Why"])
cell(ws5, 6, 1, "vocab_size (initial tokens)"); cell(ws5, 6, 2, "='7B Breakdown'!B11", NUMFMT)
why(ws5, 6, 5, "This is the size of the tokenizer's vocabulary — the fixed list of sub-word pieces (like \"ing\", \"the\", \"phot\"+\"o\") the model can ever read or write. Every input token must be one of these 32,000 IDs. Picking this number is a trade-off made BEFORE the model is trained, by whoever built the tokenizer: too few tokens and every word gets chopped into many small pieces (longer sequences, more compute per sentence); too many tokens and the embedding table below gets enormous for diminishing returns. 32,000 is a common, well-tested middle ground — this is why vocab_size is the very first number everything else depends on.")
cell(ws5, 7, 1, "hidden_size"); cell(ws5, 7, 2, "='7B Breakdown'!B6", NUMFMT)
why(ws5, 7, 5, "This is the width of the vector every single token gets turned into, and STAYS AT, through the entire model — think of it as each token's \"working memory\" size. Every matrix in the model is built to take a hidden_size-wide input and hand back a hidden_size-wide (or intermediate_size-wide, for the FFN's internal step) output. Because this one number appears in almost every formula below (often squared, in W_Q/W_K/W_V/W_O), it has an outsized effect on total parameter count — doubling hidden_size roughly QUADRUPLES the attention parameter count per layer, not just doubles it.")
cell(ws5, 8, 1, "n_layers"); cell(ws5, 8, 2, "='7B Breakdown'!B7", NUMFMT)
why(ws5, 8, 5, "How many times the token's vector gets refined. Each layer does NOT change the vector's SHAPE (still hidden_size in, hidden_size out) — it refines the CONTENT, adding one more round of \"look at other tokens, then update what I currently think.\" More layers means more sequential reasoning steps, at the direct cost of more parameters AND more sequential compute (unlike heads, layers CANNOT run in parallel with each other — layer 2 needs layer 1's finished output before it can start).")
cell(ws5, 9, 1, "n_heads"); cell(ws5, 9, 2, "='7B Breakdown'!B8", NUMFMT)
why(ws5, 9, 5, "Instead of one big attention computation over all 4,096 dimensions at once, the model splits into this many smaller, independent, PARALLEL attention computations. Why split at all, instead of one large head? A single attention pass can only learn one notion of \"what's relevant to what\" at a time. Splitting into heads lets different heads specialize — empirically, different heads DO learn different things (one might track subject-verb agreement, another might track which pronoun refers to which noun) — without any of them interfering with each other's math, since each head's slice of the vector is completely separate from every other head's slice.")
cell(ws5, 10, 1, "head_dim"); cell(ws5, 10, 2, "='7B Breakdown'!B9", NUMFMT)
why(ws5, 10, 5, "head_dim = hidden_size ÷ n_heads, exactly, with NO remainder allowed — that's a real, hard constraint on picking n_heads in the first place (it must divide hidden_size evenly, which is why you see round numbers like 32, 40, 64 for n_heads rather than arbitrary ones). This is how wide each individual head's private slice of the vector is: 128 numbers, in this model.")
cell(ws5, 11, 1, "intermediate_size"); cell(ws5, 11, 2, "='7B Breakdown'!B10", NUMFMT)
why(ws5, 11, 5, "The FFN's internal \"scratch space\" width — wider than hidden_size specifically because a wider space gives the network more room to compute new nonlinear features before compressing back down to hidden_size. This number is INDEPENDENT of hidden_size/layers/heads — it's its own separate design choice, though in practice it's picked as a consistent multiple of hidden_size (≈2.688× here — see the Notes & Caveats sheet for exactly why not the more \"obvious\" 4×).")
VOC, HID, NL, NH, HD, INT = "$B$6", "$B$7", "$B$8", "$B$9", "$B$10", "$B$11"

row = 13
section(ws5, row, 1, "STEP 1 — TOKEN EMBEDDING (initial tokens → vectors)", span=5); row += 1
header_row(ws5, row, 1, ["Stage", "Shape", "Formula", "Parameters", "Why"]); row += 1
cell(ws5, row, 1, "Initial vocabulary (distinct tokens)"); cell(ws5, row, 2, "=\"[\"&"+VOC+"&\" tokens]\""); cell(ws5, row, 3, "given, not computed"); cell(ws5, row, 4, f"={VOC}", NUMFMT)
why(ws5, row, 5, "Chosen by whoever trained the tokenizer, as its own separate step done BEFORE the model itself is ever trained — this number is an input to the model's design, not something the model computes or learns.")
row += 1
cell(ws5, row, 1, "X — one token's embedding (a lookup, NOT a matmul)"); cell(ws5, row, 2, f"=\"[1 × \"&{HID}&\"]\""); cell(ws5, row, 3, "row lookup from the table below"); cell(ws5, row, 4, "0 (no extra params — reuses the table)", None)
why(ws5, row, 5, "Technically, looking up a token's row from the embedding table IS mathematically a matrix multiply: multiply the table by a one-hot vector (all zeros except a single 1 at the token's ID) and you get exactly that token's row back — a weighted sum where every weight is 0 except one. It's implemented as a direct index/lookup instead of an actual multiply, purely because multiplying by a vector that's 31,999 zeros and one 1 would waste enormous compute for zero benefit. This is the ONE stage in the whole model that isn't dense matrix multiplication in practice, even though it's mathematically equivalent to one.", height=75)
row += 1
emb_row = row
cell(ws5, row, 1, "Embedding table W_E", bold=True, fill=INPUT_FILL); cell(ws5, row, 2, f"=\"[\"&{VOC}&\" × \"&{HID}&\"]\"", fill=INPUT_FILL); cell(ws5, row, 3, "vocab × hidden", fill=INPUT_FILL); cell(ws5, row, 4, f"={VOC}*{HID}", NUMFMT, bold=True, fill=INPUT_FILL)
why(ws5, row, 5, "One full hidden_size-wide row for EVERY possible token — literally a lookup table, 32,000 entries deep, each entry 4,096 numbers wide. Before training, every one of these 131,072,000 numbers is random noise; training is precisely the process that shapes each row into a meaningful representation of what that token tends to mean in context.")
row += 2

section(ws5, row, 1, "WORKED EXAMPLE — X × W_Q, real shape (this exact matmul happens 32 times below, once per layer)", span=5); row += 1
header_row(ws5, row, 1, ["Term", "Shape", "", "", "Why"]); row += 1
cell(ws5, row, 1, "X  (one token's hidden vector, input to this layer)"); cell(ws5, row, 2, f"=\"[1 × \"&{HID}&\"]\"")
why(ws5, row, 5, "This is the vector for ONE token, at the moment it enters THIS particular layer. On layer 1 it comes straight from the embedding table above; on layer 2 onward, it's whatever the PREVIOUS layer output. Every layer re-uses the exact same hidden_size-wide \"slot\" for both its input and its output — that shape-preservation is exactly what lets 32 of these stack back-to-back with no dimension mismatch anywhere.")
row += 1
cell(ws5, row, 1, "×  W_Q  (learned query-projection matrix)"); cell(ws5, row, 2, f"=\"[\"&{HID}&\" × \"&{HID}&\"]\"")
why(ws5, row, 5, "W_Q is a SQUARE matrix (hidden×hidden) by this architecture's design choice — it takes a hidden_size-wide input and returns a hidden_size-wide output. It doesn't HAVE to be square (some architectures shrink or grow the dimension here), but keeping it square means the \"query space\" has exactly the same capacity as the token's own representation space — no information is deliberately thrown away or padded at this step.")
row += 1
cell(ws5, row, 1, "→  Q  (this token's query vector)", bold=True, fill=TOTAL_FILL); cell(ws5, row, 2, f"=\"[1 × \"&{HID}&\"]\"", fill=TOTAL_FILL)
why(ws5, row, 5, "Matrix multiplication, concretely: each of the 4,096 output numbers in Q is the DOT PRODUCT of the full 4,096-number input X with one COLUMN of W_Q — 4,096 multiplications plus 4,095 additions, PER output number, repeated 4,096 times (once per output number). That's 4,096 × 4,096 total multiply operations for this one matrix — which is EXACTLY where the 16,777,216 parameter count below comes from: every one of those 4,096×4,096 positions in W_Q is one independent, learned number.", height=75)
row += 1
cell(ws5, row, 1, "Parameters contributed by W_Q alone", bold=True); cell(ws5, row, 3, "hidden²"); cell(ws5, row, 4, f"={HID}*{HID}", NUMFMT, bold=True)
why(ws5, row, 5, "The \"4×\" you'll see in Step 2 below is Q, K, V, and O — four SEPARATE hidden×hidden matrices per layer, each learned completely independently, each contributing its own 16,777,216 parameters. They all share the identical SHAPE but hold completely different, independently-trained VALUES: W_Q learns to build \"questions\", W_K learns to build \"answer keys\" that queries get compared against, W_V learns to build \"the actual content to retrieve\" once a match is found, and W_O learns to recombine all 32 heads' mixed-together outputs back into one coherent vector.", height=75)
row += 1
ws5.cell(row=row, column=1, value="K, V, and O use the IDENTICAL shape/formula — X×W_K=K, X×W_V=V, (mixed heads)×W_O=output — 4 matrices per layer, each [hidden×hidden]. Real trained VALUES differ per layer/matrix; the SHAPE and PARAMETER COUNT are identical every time, which is exactly why the row-by-row sum below works.").alignment = Alignment(wrap_text=True)
ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
ws5.row_dimensions[row].height = 45
row += 2

section(ws5, row, 1, "WORKED EXAMPLE — the FFN, and why it's 3 matrices (not 2, not 4)", span=5); row += 1
header_row(ws5, row, 1, ["Term", "Shape", "", "", "Why"]); row += 1
cell(ws5, row, 1, "h  (this layer's post-attention vector, FFN's input)"); cell(ws5, row, 2, f"=\"[1 × \"&{HID}&\"]\"")
why(ws5, row, 5, "The FFN runs AFTER attention finishes and its residual has been added — same hidden_size shape as X was for attention, for the identical reason: every stage in a layer must take hidden_size in and hand back hidden_size out, so stages (and whole layers) can chain without a dimension mismatch.", height=55)
row += 1
cell(ws5, row, 1, "×  W_gate  (learned \"how much to let through\" matrix)"); cell(ws5, row, 2, f"=\"[\"&{HID}&\" × \"&{INT}&\"]\"")
why(ws5, row, 5, "First of the FFN's two PARALLEL up-projections. W_gate expands h from hidden_size (4,096) up into the FFN's wider intermediate_size (11,008) \"scratch space\" — then gets passed through a SiLU/Swish nonlinearity, producing a per-feature signal roughly meaning \"how open is this gate.\"", height=70)
row += 1
cell(ws5, row, 1, "×  W_up  (learned \"what content\" matrix, computed in parallel)"); cell(ws5, row, 2, f"=\"[\"&{HID}&\" × \"&{INT}&\"]\"")
why(ws5, row, 5, "Second up-projection, SAME shape as W_gate but a completely separate, independently-learned matrix — no nonlinearity applied here, just a linear expansion. This is the actual \"content\" signal that the gate above will modulate.", height=60)
row += 1
cell(ws5, row, 1, "→  SiLU(gate) ⊙ up  (elementwise multiply — THIS is why 2 matrices, not 1)", bold=True, fill=TOTAL_FILL); cell(ws5, row, 2, f"=\"[1 × \"&{INT}&\"]\"", fill=TOTAL_FILL)
why(ws5, row, 5, "The \"⊙\" is an ELEMENTWISE multiply, not a matmul: each of the 11,008 numbers in the gated result is simply (SiLU(gate)[i] × up[i]) — same-position numbers multiplied together, nothing else. This is a GATING mechanism: the gate branch (put through SiLU, squashed toward 0 or passed through) decides how much of the up branch's content survives into the FFN's output, feature by feature. A single ordinary matmul-then-nonlinearity (what a classic 2-matrix GELU FFN does) has no way to let one learned signal modulate another this way — that extra expressive power is specifically what costs the 3rd matrix.", height=110)
row += 1
cell(ws5, row, 1, "×  W_down  (compress back down)"); cell(ws5, row, 2, f"=\"[\"&{INT}&\" × \"&{HID}&\"]\"")
why(ws5, row, 5, "The third matrix: maps the 11,008-wide gated result back down to hidden_size (4,096) so it can be added back via the residual connection, same as attention's output was. This is the mirror image of gate/up's expansion — intermediate_size in, hidden_size out.", height=60)
row += 1
cell(ws5, row, 1, "Parameters: gate + up + down, all three", bold=True); cell(ws5, row, 3, "3 × hidden × intermediate"); cell(ws5, row, 4, f"=3*{HID}*{INT}", NUMFMT, bold=True)
why(ws5, row, 5, "THE ANSWER TO \"WHY 3×\": it is not a multiplier on size, it is literally a COUNT of three separate, independently-learned [hidden×intermediate]-shaped matrices — W_gate, W_up, and W_down — each contributing hidden×intermediate parameters. A classic pre-SwiGLU FFN only needs 2 matrices (one up-projection through a plain nonlinearity, one down-projection) because it has no gating branch to compute separately; SwiGLU's elementwise gate-times-up step (previous row) is exactly the mechanism that requires the extra, 3rd matrix. This is also why intermediate_size is ≈2.688×hidden rather than the classic 4×hidden — see the Notes & Caveats sheet: 3 matrices at a smaller multiplier lands at roughly the SAME total parameter budget as 2 matrices at 4×.", height=120)
row += 2

section(ws5, row, 1, "STEP 2 — ALL 32 LAYERS, ENUMERATED (every layer listed, nothing compressed as ×32)", span=5); row += 1
header_row(ws5, row, 1, ["Layer #", "Attention (Q+K+V+O), 4×hidden²", "FFN (gate+up+down), 3×hidden×intermediate", "Layer total", "Why"]); row += 1
layer_first_row = row
for L in range(1, 33):
    cell(ws5, row, 1, f"Layer {L}")
    cell(ws5, row, 2, f"=4*{HID}*{HID}", NUMFMT)
    cell(ws5, row, 3, f"=3*{HID}*{INT}", NUMFMT)
    cell(ws5, row, 4, f"=B{row}+C{row}+2*{HID}", NUMFMT)
    if L == 1:
        why(ws5, row, 5, "This row's two numbers (67,108,864 for attention, 135,266,304 for FFN) are the exact formulas worked out above, applied to THIS layer's own, independently-trained W_Q/W_K/W_V/W_O/W_gate/W_up/W_down matrices. Every one of the 32 layers has this IDENTICAL shape and IDENTICAL parameter count — but each layer's actual matrix VALUES are learned completely independently during training. That's why the total is trivially \"32 × one_layer\" arithmetically, while the model's actual behavior is NOT just repeating the same transformation 32 times — layer 2 is a different learned function from layer 1, applied to whatever layer 1 handed it.", height=90)
    else:
        why(ws5, row, 5, f"Same shape/formula as Layer 1 — see the explanation there. Only the trained VALUES differ between layers, never the shape or the parameter count.", height=20)
    row += 1
layer_last_row = row - 1
cell(ws5, row, 1, "ALL 32 LAYERS — TOTAL", bold=True, fill=TOTAL_FILL)
cell(ws5, row, 4, f"=SUM(D{layer_first_row}:D{layer_last_row})", NUMFMT, bold=True, fill=TOTAL_FILL)
why(ws5, row, 5, "A plain SUM across all 32 rows above — this works cleanly BECAUSE every layer's parameters are its own, entirely separate set of matrices; nothing here is shared or reused across layers (unlike the embedding table, which IS reused — see Step 1).", height=45)
all_layers_row = row
row += 2

section(ws5, row, 1, "STEP 3 — INSIDE ONE LAYER'S ATTENTION: ALL 32 HEADS, ENUMERATED", span=5); row += 1
ws5.cell(row=row, column=1, value="hidden_size splits into n_heads × head_dim (4,096 = 32 × 128). These 32 rows are a VIEW into the SAME W_Q/W_K/W_V matrices already counted in Step 2 above — splitting into heads is a reshape, NOT additional parameters. This section explains SHAPE; Step 2 already counted the PARAMETERS. Adding these to the total would double-count.").alignment = Alignment(wrap_text=True)
ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
ws5.row_dimensions[row].height = 45
row += 1
header_row(ws5, row, 1, ["Head #", "Columns of the full 4096-dim Q/K/V it uses", "Q_head / K_head / V_head shape", "Params added by this head", "Why"]); row += 1
head_first_row = row
for h in range(1, 33):
    lo = f"(({h}-1)*{HD}+1)"
    hi = f"({h}*{HD})"
    cell(ws5, row, 1, f"Head {h}")
    cell(ws5, row, 2, f"=\"[\"&{lo}&\" : \"&{hi}&\"]\"")
    cell(ws5, row, 3, f"=\"[1 × \"&{HD}&\"]\"")
    cell(ws5, row, 4, "0 — reshape only, see note above")
    if h == 1:
        why(ws5, row, 5, "hidden_size=4,096 gets sliced into 32 contiguous chunks of 128 numbers each — head 1 owns dimensions 1–128 of Q, K, and V (its own private slice, touched by no other head). Each head then computes its OWN attention independently: softmax(Q_head · K_headᵀ ÷ √128) · V_head — completely separate math from what the other 31 heads are doing with their own 128-number slices, running in parallel. After all 32 finish, their 32 separate 128-number outputs get concatenated back into one 4,096-number vector (32×128=4,096 exactly — no gaps, no overlap) before passing through W_O.", height=105)
    else:
        why(ws5, row, 5, "Same mechanism as Head 1 — just the next contiguous 128-column slice of Q, K, and V.", height=20)
    row += 1
head_last_row = row - 1
cell(ws5, row, 1, "Check: heads × head_dim = hidden_size?", bold=True)
cell(ws5, row, 2, f"=({head_last_row-head_first_row+1})*{HD}", NUMFMT)
cell(ws5, row, 3, f"=IF(B{row}={HID},\"✓ matches hidden_size exactly\",\"✗ mismatch\")")
why(ws5, row, 5, "A structural sanity check, not a coincidence: n_heads × head_dim MUST equal hidden_size exactly, because head_dim was DEFINED as hidden_size ÷ n_heads back in the Model Dimensions section. If this ever showed a mismatch, it would mean the model config itself was invalid.", height=60)
row += 2

section(ws5, row, 1, "STEP 4 — FINAL STAGES", span=5); row += 1
header_row(ws5, row, 1, ["Stage", "Shape", "Formula", "Parameters", "Why"]); row += 1
final_norm_row = row
cell(ws5, row, 1, "Final Norm"); cell(ws5, row, 2, f"=\"[\"&{HID}&\"]\""); cell(ws5, row, 3, "hidden"); cell(ws5, row, 4, f"={HID}", NUMFMT)
why(ws5, row, 5, "Right before turning the vector into vocabulary scores, the model rescales it (RMSNorm: divide by the vector's own root-mean-square magnitude, then multiply by a learned per-dimension scale). Without this, the vector's magnitude could drift arbitrarily large or small after 32 stacked layers of residual additions, making the final projection numerically unstable. Only 4,096 learned numbers (one scale factor per dimension) — tiny next to everything else, but structurally necessary for training to work at all.", height=90)
row += 1
output_head_row = row
cell(ws5, row, 1, "Output Head (LM head, untied)"); cell(ws5, row, 2, f"=\"[\"&{HID}&\" × \"&{VOC}&\"]\""); cell(ws5, row, 3, "hidden × vocab"); cell(ws5, row, 4, f"={HID}*{VOC}", NUMFMT)
why(ws5, row, 5, "Maps the final hidden_size-wide vector to ONE SCORE PER VOCABULARY WORD (32,000 of them) — the exact same shape as the embedding table, just used in the opposite direction (hidden→vocab here, vs. vocab→hidden there). Llama trains this as a SEPARATE matrix from the input embedding (\"untied\") rather than reusing the embedding table transposed — costing an extra 131,072,000 parameters, but letting the model learn different things for \"how do I read a word\" versus \"how do I predict a word,\" which aren't necessarily symmetric tasks.", height=90)
row += 2

section(ws5, row, 1, "STEP 5 — GRAND TOTAL (the addition, running total — this IS how 7 billion is built)", span=5); row += 1
header_row(ws5, row, 1, ["Component", "Parameters", "Running total", "", "Why"]); row += 1
r1 = row
cell(ws5, row, 1, "Token Embedding  (Step 1)"); cell(ws5, row, 2, f"=D{emb_row}", NUMFMT); cell(ws5, row, 3, "=B"+str(row), NUMFMT)
why(ws5, row, 5, "The very first bucket added — every token's starting representation, before it ever reaches a single transformer layer.", height=35)
row += 1
r2 = row
cell(ws5, row, 1, "All 32 Layers  (Step 2)"); cell(ws5, row, 2, f"=D{all_layers_row}", NUMFMT); cell(ws5, row, 3, f"=C{row-1}+B{row}", NUMFMT)
why(ws5, row, 5, "By far the largest bucket (96.1% of the total) — this is where essentially all of the model's \"thinking\" capacity lives: every layer's attention (who should I pay attention to) and FFN (what should I conclude from that) parameters, all 32 of them summed.", height=55)
row += 1
r3 = row
cell(ws5, row, 1, "Final Norm  (Step 4)"); cell(ws5, row, 2, f"=D{final_norm_row}", NUMFMT); cell(ws5, row, 3, f"=C{row-1}+B{row}", NUMFMT)
why(ws5, row, 5, "Tiny (4,096 out of 6.7 billion) but structurally necessary — see the Step 4 explanation above.", height=35)
row += 1
r4 = row
cell(ws5, row, 1, "Output Head  (Step 4)"); cell(ws5, row, 2, f"=D{output_head_row}", NUMFMT); cell(ws5, row, 3, f"=C{row-1}+B{row}", NUMFMT)
why(ws5, row, 5, "The final translation step, from \"hidden thought\" back into vocabulary scores — the last bucket added before the running total equals the model's full parameter count.", height=45)
row += 1
cell(ws5, row, 1, "TOTAL TRAINABLE PARAMETERS", bold=True, fill=TOTAL_FILL)
cell(ws5, row, 2, f"=SUM(B{r1}:B{r4})", NUMFMT, bold=True, fill=TOTAL_FILL)
cell(ws5, row, 3, f"=C{r4}", NUMFMT, bold=True, fill=TOTAL_FILL)
cell(ws5, row, 4, "← this is how 7,000,000,000 comes from real shapes", fill=TOTAL_FILL)
why(ws5, row, 5, "This addition — embedding + all 32 layers + final norm + output head — has NO overlap and NO double-counting: every single parameter in the real model belongs to EXACTLY ONE of these four buckets (the 32-heads section above is explicitly excluded from this sum, since it's a reshape view already counted inside \"All 32 Layers\"). That's why a plain SUM here gives the exact right answer rather than an approximation — nothing was estimated anywhere in this sheet.", height=75)
row += 1
cell(ws5, row, 1, "  in billions"); cell(ws5, row, 2, f"=C{row-1}/1000000000"); ws5.cell(row=row, column=2).number_format = "0.000"
row += 1
cell(ws5, row, 1, "  cross-check against '7B Breakdown' sheet"); cell(ws5, row, 2, f"=IF(C{row-2}='7B Breakdown'!C26,\"✓ matches exactly\",\"✗ mismatch\")")
why(ws5, row, 5, "This sheet built the total completely independently (real shapes, layer by layer, head by head) from Sheet 1's more compressed \"×32\" version — if the two ever disagreed, it would mean one of the two formula sets has a bug. They agree.", height=55)

ws5.freeze_panes = "A5"


out_path = r"D:\nvidia\llama2-parameter-breakdown.xlsx"
wb.save(out_path)
print("saved:", out_path)

# ============================================================
# SELF-VERIFICATION: label/explanation text that accidentally starts
# with "=" gets parsed by Excel as a real formula and breaks the file
# -- this has happened 3 times while writing this script by hand.
# Rather than guess which cells SHOULD be formulas, actually load the
# saved file with a real formula engine and check every cell for real
# errors, every time this script runs -- fail loudly if anything's wrong
# instead of silently shipping a broken workbook.
# ============================================================
import sys
try:
    import formulas
    import numpy as np
    xl = formulas.ExcelModel().loads(out_path).finish()
    sol = xl.calculate()
    errors = []
    checked = 0
    for k, v in sol.items():
        try:
            arr = np.array(v.value).flatten()
        except Exception:
            continue
        for item in arr:
            checked += 1
            if str(item).startswith("#"):
                errors.append((k, str(item)))
    if errors:
        print(f"SELF-CHECK FAILED -- {len(errors)} error cell(s) out of {checked} checked:")
        for k, s in errors[:20]:
            print(f"  {k} -> {s}")
        sys.exit(1)
    print(f"self-check passed: {checked} cells evaluated, 0 errors")
except ImportError:
    print("self-check skipped: 'formulas' package not installed (pip install formulas)")
except Exception as e:
    print(f"SELF-CHECK FAILED to even load/evaluate the workbook: {e!r}")
    sys.exit(1)
